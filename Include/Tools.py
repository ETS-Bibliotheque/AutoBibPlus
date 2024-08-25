# © 2023 Benjamin Lepourtois <benjamin.lepourtois@gmail.com>
# © 2024 Adji Toure <adji.toure.dev@gmail.com>
# Copyright: All rights reserved.
# See the license attached to the root of the project.

"""
Projet d'Automatisation de Rapports d'Analyses Bibliométriques:

Ce programme fait partie intégrante du projet de conception et développement d'outils automatisés pour la réalisation de rapports d'analyses bibliométriques.

Contexte:

  ● Stage de 12 semaines sur l'été 2023 (12 juin au 1er septembre) dans l'École de Technologie Supérieure, Montréal, Canada
  ● Mission principale:
Développer des outils permettant l'automatisation de certaines étapes de production de rapports d'analyses bibliométriques 
destinés à aider les chercheurs et chercheuses dans la planification de la mesure de l'impact de leurs contributions scientifiques.

Approche choisie:

Nous avons choisi d'utiliser un script Python pour gérer toute l'automatisation des rapports.
  ● Extraction des données: par les API des différentes plateformes utilisées (Scopus et SciVal) à l'aide de la bibliothèque publique "pybliometrics"
  ● Traitement des données: en Python à l'aide de la bibliothèque "pandas"
  ● Interface Homme-Machine: en QT avec une interface très simpliste basée sur une boîte de dialogue
  ● Exportation des données: en Python à l'aide de la bibliothèque "pywin32" vers un fichier "Workbook" MacroExcel (.xlsm)
  ● Mise en forme Excel: avec des routines VBA appelées par le script Python
  ● Réalisation du rapport Word: avec des routines VBA, appelées par le script Python, qui exportent les données et les graphiques réalisés sur un document Word
"""


""" 
Projet repris par Adji Touré
Contexte :
  ● Stage de 12 semaines sur l'été 2024 (3 juin au 23 août) à l'École de Technologie Supérieure, Montréal, Canada.

Mission principale : 
  ● Continuer le développement de l'outil en intégrant la production de rapports de colloaborations entre deux entités.

Approche adoptée et tâches réalisées:
  ● Observation de la méthodologie existante : Analyse des processus et méthodes actuels utilisés pour la réalisation des rapports bibliométriques.
  ● Participation aux échanges avec les fournisseurs : Interaction avec les partenaires externes pour mieux comprendre les outils et services disponibles.
  ● Évaluation de différentes approches : Identification et test de la faisabilité de diverses approches pour le projet, analyse des avantages et inconvénients de chacune, et présentation de recommandations à l'équipe pour sélectionner l'approche optimale.
  ● Maîtrise des bibliothèques et APIs : Étude approfondie de la documentation sur l'exploitation des APIs, notamment celles d'Elsevier, et acquisition de compétences en utilisant des bibliothèques comme pybliometrics pour l'extraction de données et pandas pour leur traitement.
  ● Développement de scripts pour la détection de noms : Création d'un script Python pour détecter et identifier les noms des professeurs de l'établissement en utilisant des techniques de correspondance approximative.
  ● Automatisation des rapports : Développement de scripts Python pour :
        Extraction des données bibliométriques via l'API Scopus.
        Traitement et calcul des indicateurs requis.
        Exportation des données vers Excel et intégration des graphiques dans des rapports Word via des routines VBA.
        Amélioration de l'Interface Homme-Machine (IHM) : Développement d'une nouvelle branche pour l'IHM en Qt.
"""

import os, unicodedata, win32gui, time, re
import pandas as pd
import json
from unidecode import unidecode
from fuzzywuzzy import fuzz
import win32com.client as win32
from datetime import datetime
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from xlsxwriter import Workbook

# Importations locales
from .pybliometrics.scopus.abstract_citation import CitationOverview
from .pybliometrics.scopus.author_retrieval import AuthorRetrieval
from .pybliometrics.scopus.author_search import AuthorSearch
from .pybliometrics.scopus.abstract_retrieval import AbstractRetrieval
from .pybliometrics.scival.author_lookup import AuthorLookup
from .pybliometrics.scival.institution_lookup import InstitutionLookup
from .pybliometrics.utils.startup import DOCS_PATH
from .pybliometrics.scopus.affiliation_retrieval import AffiliationRetrieval
from .pybliometrics.scopus.affiliation_search import AffiliationSearch
from Include.pybliometrics.scopus.scopus_search import ScopusSearch

# Pour utiliser la console de l'IHM
from PySide6.QtWidgets import QPlainTextEdit

# Appliquer une feuille de style CSS pour le texte en couleur
text_style_warning = '"color: #D35230"'
text_style_question = '"color: #0C5E31"'

# Dictionnaires des traductions souhaitées
trad_en2fr = {
    'Article': 'Article',
    'Review': 'Article de synthèse',
    'Conference Paper': 'Conférence',
    'Conference Review': 'Synthèse de conférence',
    'Data Paper': 'Article de données',
    'Editorial': 'Éditorial',
    'Book': 'Livre',
    'Book Chapter': 'Chapitre',
    'Erratum': 'Erratum',
    'Note': 'Commentaire',
    'Letter': "Lettre d'opinion",
    'Short Survey': 'Enquête',
    'Retracted': 'Rétractation',
    '∅': '∅'
}
countries_dict = {
    "Afghanistan": "Afghanistan",
    "Afrique du Sud": "South Africa",
    "Albanie": "Albania",
    "Algérie": "Algeria",
    "Allemagne": "Germany",
    "Andorre": "Andorra",
    "Angola": "Angola",
    "Antigua-et-Barbuda": "Antigua and Barbuda",
    "Arabie Saoudite": "Saudi Arabia",
    "Argentine": "Argentina",
    "Arménie": "Armenia",
    "Australie": "Australia",
    "Autriche": "Austria",
    "Azerbaïdjan": "Azerbaijan",
    "Bahamas": "Bahamas",
    "Bahreïn": "Bahrain",
    "Bangladesh": "Bangladesh",
    "Barbade": "Barbados",
    "Belgique": "Belgium",
    "Belize": "Belize",
    "Bénin": "Benin",
    "Bhoutan": "Bhutan",
    "Biélorussie": "Belarus",
    "Birmanie": "Myanmar",
    "Bolivie": "Bolivia",
    "Bosnie-Herzégovine": "Bosnia and Herzegovina",
    "Botswana": "Botswana",
    "Brésil": "Brazil",
    "Brunei": "Brunei",
    "Bulgarie": "Bulgaria",
    "Burkina Faso": "Burkina Faso",
    "Burundi": "Burundi",
    "Cambodge": "Cambodia",
    "Cameroun": "Cameroon",
    "Canada": "Canada",
    "Cap-Vert": "Cape Verde",
    "Chili": "Chile",
    "Chine": "China",
    "Chypre": "Cyprus",
    "Colombie": "Colombia",
    "Comores": "Comoros",
    "Congo-Brazzaville": "Republic of the Congo",
    "Congo-Kinshasa": "Democratic Republic of the Congo",
    "Corée du Nord": "North Korea",
    "Corée du Sud": "South Korea",
    "Costa Rica": "Costa Rica",
    "Côte d'Ivoire": "Ivory Coast",
    "Croatie": "Croatia",
    "Cuba": "Cuba",
    "Danemark": "Denmark",
    "Djibouti": "Djibouti",
    "Dominique": "Dominica",
    "Égypte": "Egypt",
    "Émirats Arabes Unis": "United Arab Emirates",
    "Équateur": "Ecuador",
    "Érythrée": "Eritrea",
    "Espagne": "Spain",
    "Estonie": "Estonia",
    "États-Unis": "United States",
    "Éthiopie": "Ethiopia",
    "Fidji": "Fiji",
    "Finlande": "Finland",
    "France": "France",
    "Gabon": "Gabon",
    "Gambie": "Gambia",
    "Géorgie": "Georgia",
    "Ghana": "Ghana",
    "Grèce": "Greece",
    "Grenade": "Grenada",
    "Guatemala": "Guatemala",
    "Guinée": "Guinea",
    "Guinée-Bissau": "Guinea-Bissau",
    "Guinée équatoriale": "Equatorial Guinea",
    "Guyana": "Guyana",
    "Haïti": "Haiti",
    "Honduras": "Honduras",
    "Hongrie": "Hungary",
    "Inde": "India",
    "Indonésie": "Indonesia",
    "Irak": "Iraq",
    "Iran": "Iran",
    "Irlande": "Ireland",
    "Islande": "Iceland",
    "Israël": "Israel",
    "Italie": "Italy",
    "Jamaïque": "Jamaica",
    "Japon": "Japan",
    "Jordanie": "Jordan",
    "Kazakhstan": "Kazakhstan",
    "Kenya": "Kenya",
    "Kirghizistan": "Kyrgyzstan",
    "Kiribati": "Kiribati",
    "Koweït": "Kuwait",
    "Laos": "Laos",
    "Lesotho": "Lesotho",
    "Lettonie": "Latvia",
    "Liban": "Lebanon",
    "Liberia": "Liberia",
    "Libye": "Libya",
    "Liechtenstein": "Liechtenstein",
    "Lituanie": "Lithuania",
    "Luxembourg": "Luxembourg",
    "Macédoine": "North Macedonia",
    "Madagascar": "Madagascar",
    "Malaisie": "Malaysia",
    "Malawi": "Malawi",
    "Maldives": "Maldives",
    "Mali": "Mali",
    "Malte": "Malta",
    "Maroc": "Morocco",
    "Marshall": "Marshall Islands",
    "Maurice": "Mauritius",
    "Mauritanie": "Mauritania",
    "Mexique": "Mexico",
    "Micronésie": "Micronesia",
    "Moldavie": "Moldova",
    "Monaco": "Monaco",
    "Mongolie": "Mongolia",
    "Monténégro": "Montenegro",
    "Mozambique": "Mozambique",
    "Namibie": "Namibia",
    "Nauru": "Nauru",
    "Népal": "Nepal",
    "Nicaragua": "Nicaragua",
    "Niger": "Niger",
    "Nigéria": "Nigeria",
    "Niue": "Niue",
    "Norvège": "Norway",
    "Nouvelle-Zélande": "New Zealand",
    "Oman": "Oman",
    "Ouganda": "Uganda",
    "Ouzbékistan": "Uzbekistan",
    "Pakistan": "Pakistan",
    "Palaos": "Palau",
    "Palestine": "Palestine",
    "Panama": "Panama",
    "Papouasie-Nouvelle-Guinée": "Papua New Guinea",
    "Paraguay": "Paraguay",
    "Pays-Bas": "Netherlands",
    "Pérou": "Peru",
    "Philippines": "Philippines",
    "Pologne": "Poland",
    "Portugal": "Portugal",
    "Qatar": "Qatar",
    "Roumanie": "Romania",
    "Royaume-Uni": "United Kingdom",
    "Russie": "Russia",
    "Rwanda": "Rwanda",
    "Saint-Kitts-et-Nevis": "Saint Kitts and Nevis",
    "Saint-Vincent-et-les-Grenadines": "Saint Vincent and the Grenadines",
    "Sainte-Lucie": "Saint Lucia",
    "Salomon": "Solomon Islands",
    "Salvador": "El Salvador",
    "Samoa": "Samoa",
    "Sao Tomé-et-Principe": "Sao Tome and Principe",
    "Sénégal": "Senegal",
    "Serbie": "Serbia",
    "Seychelles": "Seychelles",
    "Sierra Leone": "Sierra Leone",
    "Singapour": "Singapore",
    "Slovaquie": "Slovakia",
    "Slovénie": "Slovenia",
    "Somalie": "Somalia",
    "Soudan": "Sudan",
    "Soudan du Sud": "South Sudan",
    "Sri Lanka": "Sri Lanka",
    "Suède": "Sweden",
    "Suisse": "Switzerland",
    "Suriname": "Suriname",
    "Swaziland": "Eswatini",
    "Syrie": "Syria",
    "Tadjikistan": "Tajikistan",
    "Tanzanie": "Tanzania",
    "Tchad": "Chad",
    "Tchéquie": "Czech Republic",
    "Thaïlande": "Thailand",
    "Timor oriental": "East Timor",
    "Togo": "Togo",
    "Tonga": "Tonga",
    "Trinité-et-Tobago": "Trinidad and Tobago",
    "Tunisie": "Tunisia",
    "Turkménistan": "Turkmenistan",
    "Turquie": "Turkey",
    "Tuvalu": "Tuvalu",
    "Ukraine": "Ukraine",
    "Uruguay": "Uruguay",
    "Vanuatu": "Vanuatu",
    "Vatican": "Vatican City",
    "Venezuela": "Venezuela",
    "Vietnam": "Vietnam",
    "Yémen": "Yemen",
    "Zambie": "Zambia",
    "Zimbabwe": "Zimbabwe"
}

def remove_accents(input_str : str):
    nfkd_form = unicodedata.normalize('NFKD', input_str)
    return "".join([c for c in nfkd_form if not unicodedata.combining(c)])

def get_country_in_english(country_name : str):
    country_name_normalized = remove_accents(country_name).lower()
    for key in countries_dict.keys():
        if remove_accents(key).lower().replace(' ', '') == country_name_normalized:
            return countries_dict[key]
    return 'NULL'
def get_country_in_french(country_name : str):
    country_name_normalized = remove_accents(country_name).lower()
    for key in countries_dict.keys():
        if remove_accents(key).lower().replace(' ', '') == country_name_normalized:
            return key
    return 'NULL'
def get_country_for_request(country_name : str):
    country_name = get_country_in_english(country_name)
    if ' ' in country_name:
        country_name = country_name.replace(' ', ' AND ')
        return country_name
    return country_name

# Inverser le dictionnaire en échangeant les clés et les valeurs
trad_fr2en = {v: k for k, v in trad_en2fr.items()}

# Fonction qui retourne l'incrément de la variable d'état en fonction du nombre de résultats pour un nom et un prénom de la personne
def homonyme(resultatRecherche: AuthorSearch, console: QPlainTextEdit, window_width: int):
    if not resultatRecherche.get_results_size():
        console.append('<p style={}>! Aucun résultat</p>'.format(text_style_warning))
        console.append('')
        console.append('<p style={}>● Veuillez entrer le nom et le prénom de la personne [respectivement avec virgule comme séparateur] :</p>'.format(text_style_question))
        return 0
    else:
        # Affiche les résultats de manière organisée
        pd.set_option('display.max_columns', None)
        df = pd.DataFrame(resultatRecherche.authors)
        df.index.name = 'Index'
        df = df.drop(df.columns[[0, 1, 3, 7]], axis=1) # Supprime les colonnes de données qui ne nous intéressent pas
        df.columns = ['Nom', 'Prénom', 'Affiliation', 'Nb docs', 'Ville', 'Pays', 'Domaine(s) de recherche'] # Renomme alors celles qui nous intéressent
        console.append('')
        console.append('<p style="text-decoration: underline; color: black;">Personne.s trouvée.s :</p>')
        console.append(df.to_string(index=True, col_space=0, line_width=window_width))

        # Permet de savoir si l'utilisateur doit choisir une personne dans une liste
        if len(resultatRecherche.authors) > 1:
            console.append('')
            console.append("<p style={}>● Quelle personne choisissez-vous? [Entrez le numéro de l'index]</p>".format(text_style_question))
            return 1
    return 2

# Fonction utilitaire pour la fonction "selection_homonyme"
def _is_valid_integer(value, max_value):
    if value.isdigit():
        return 0 <= int(value) < max_value
    return False
    
# Fonction qui retourne vrai si les index rentrés sont valides
def selection_homonyme(choix: str, s: AuthorSearch, console: QPlainTextEdit):
    # Vérifier si la valeur entrée est un entier et compris dans range de personnes trouvées
    if _is_valid_integer(choix, len(s.authors)):
        return True
    else:
        console.append('<p style={}>! Veuillez entrer un index du tableau valide</p>'.format(text_style_warning))
        console.append('')
        console.append("<p style={}>● Quelle personne choisissez-vous? [Entrez le numéro de l'index]</p>".format(text_style_question))
        return False


# Fonction qui retourne l'EID tronqué et surtout l'instance de AuthorRetrieval sur la personne sélectionnée
def retrieval(choix: int, s: AuthorSearch, console: QPlainTextEdit):
    # Récupération de l'identifier de l'eid en fonction de la personne sélectionné
    author_eid = s.authors[choix].eid
    author_eid = author_eid.split("s2.0-")[-1] # récupère le 2ème élément créé avec le split (donc eid)

    # Affiche un résumé sur la personne sélectionnée
    au_retrieval = AuthorRetrieval(author_eid, refresh=True)
    sum = str(au_retrieval)

    console.append('')
    console.append('<p style="text-decoration: underline;">Résumé de la personne sélectionnée :</p>')
    console.append('<p style="font-weight: bold;">{}</p>'.format(sum))
    console.append('\n')

    return author_eid, au_retrieval

# Fonction qui retourne le scopus ID et l'instance de AffiliationRetrieval sur l'entité selectionnée
def affRetrieval(choix: int, s: AffiliationSearch, console: QPlainTextEdit):
    # Récupération de l'identifier de l'eid en fonction de la personne sélectionné
    affiliation_eid = s.affiliations[choix].eid
    affiliation_eid = affiliation_eid.split("s2.0-")[-1] # récupère le 2ème élément créé avec le split (donc eid)

    # Affiche un résumé sur l'entité sélectionnée
    aff_retrieval = AuthorRetrieval(affiliation_eid, refresh=True)

    return affiliation_eid, aff_retrieval

# Fonction qui retourne un DataFrame sur les types de documents avec leur nombre en fonction de la personne sélectionnée
def tous_les_docs_chercheur(au_retrieval: AuthorRetrieval, console: QPlainTextEdit):
    # Récupère tous les documents publiés de la personne et les stock dans un DataFrame
    docs = pd.DataFrame(au_retrieval.get_documents(refresh=10))

    # Afficher les valeurs uniques dans la colonne 'subtypeDescription'
    list_val = docs['subtypeDescription'].unique()

    # Calcul le nombre total de document du personne
    total = sum(len(docs[docs['subtypeDescription'] == val]) for val in list_val)

    # Compter les occurrences de chaque valeur
    value_counts = docs['subtypeDescription'].value_counts()

    # Créer un DataFrame avec index de ref et données={type_de_document, value_counts}, renommage des colonnes de données
    df = pd.DataFrame({'count': value_counts})
    df = df.reset_index()
    df.columns = ['Type de documents', 'Nombre']
    df.index.name = 'Index'

    # Appliquer la traduction aux types de documents
    df['Type de documents'] = df['Type de documents'].map(trad_en2fr)

    # Affichage
    console.append('<p style="text-decoration: underline; color: black;">Nb de documents en fonction de leur type :</p>')
    console.append(df.to_string(index=True, col_space=0, line_width=200)) # .to_string().encode('utf-8')
    console.append('<p style="font-weight: bold;">Total: {}</p>'.format(str(total)))

    return df

# Fonction qui retourne un DataFrame contenant toutes les collaborations d'une entité
def tous_les_docs_entite(aff_retrieval: AffiliationRetrieval):
    # Récupère tous les documents publiés de la personne et les stock dans un DataFrame
    docs = pd.DataFrame(aff_retrieval.__str__())
    return docs

# Fonction qui retourne vrai si la sélection des types est correcte
def selection_types_de_documents(selected_types: list, len_df: int, console: QPlainTextEdit):
    tout_valide = True

    # Pour chaque type de docs sélectionné vérifier si l'index est valide si seulement un n'est pas valide alors la fonction retournera faux
    for type_index in selected_types:
        # Vérifier si l'index est valide
        if not (type_index.isdigit() and int(type_index) < len_df):
            console.append('<p style={}>! Index non valide : {}</p>'.format(text_style_warning, type_index))
            tout_valide = False
        elif selected_types.count(str(int(type_index))) > 1:
            console.append('<p style={}>! Doublon trouvé : {}</p>'.format(text_style_warning, type_index))
            tout_valide = False        
    
    return tout_valide

# Fonction qui retourne les listes de : du nombre de documents par année avec prise en compte des types de docs sélectionnés, 
# des eids de tous les documents des types sélectionnés, ainsi que les années de carrière de la personne
def donnees_documents_graph_citations(au_retrieval: AuthorRetrieval, selected_types: list, df: pd.DataFrame, console: QPlainTextEdit):
    # Créé un DataFrame avec toutes les données sur tous les documents de la personne sélectionnée
    docs = pd.DataFrame(au_retrieval.get_documents(refresh=10))

    # Trie par année de publication des documents
    draft_list = docs['coverDate'].str[:4].sort_values() # pour draft list

    # Liste tous les types sélectionnés avec la retraduction en anglais (pour les matchs juste après)
    selected_types = [trad_fr2en[doc] for doc in [df['Type de documents'].loc[int(type_index)] for type_index in selected_types]]

    # Filtrer les documents en fonction des types sélectionnés
    filtered_docs = [len(docs[docs['subtypeDescription'] == type_info]) for type_info in selected_types]

    # Créer un DataFrame avec index=list_val et données=value_counts
    df2 = pd.DataFrame({'Type de documents': selected_types,
                        'Nombre': filtered_docs})

    # Affichage des traductions des types de docs du df2 (les sélectionnés) les uns après les autres séparés par une virgule
    selection_string = ', '.join([trad_en2fr[doc] for doc in df2['Type de documents'].tolist()])
    console.append('')
    console.append('<a style="font-weight: bold;">Votre sélection : </a>' + selection_string)
    console.append('\n')

    # Filtrer les documents en fonction des types sélectionnés et créer une nouvelle colonne 'Année'
    filtered_docs2 = docs.loc[docs['subtypeDescription'].isin(selected_types), ['subtypeDescription', 'coverDate']].copy()
    filtered_docs2['Année'] = filtered_docs2['coverDate'].str[:4]

    # Grouper les documents filtrés par année et effectuer la somme des articles par année
    counts_par_annee = filtered_docs2.groupby('Année').size()

    # Créer un DataFrame à partir des comptages par année et affichage avec transposition
    df_final = pd.DataFrame({'Nombre': counts_par_annee})

    # Créé la liste utilisée pour l'exportation des données dans Excel
    liste = df_final.reset_index().values.tolist()
    first_year = draft_list.values[0]
    total_annees = datetime.now().year - int(first_year) + 2

    # Liste de toutes les années de la personne
    years = [int(first_year) + i for i in range(total_annees)]

    # Créé la liste finale avec le nombre total de citations par année
    final_list = [0] * total_annees
    index_liste = 0
    for i in range(int(first_year), int(first_year) + total_annees):
        if index_liste < len(liste) and int(liste[index_liste][0]) == i:
            final_list[i - int(first_year)] = liste[index_liste][1]
            index_liste += 1

    # Ajoute le total de cette liste à la fin de la liste (écrasement/overwriting)
    final_list.append(sum(final_list))

    # Créé une liste de tous les eids des documents qui sont des types sélectionnés
    eids_list = docs.loc[docs['subtypeDescription'].isin(selected_types), ['eid']].copy()
    eids_list = eids_list['eid'].tolist()

    return final_list, eids_list, years

# Fonction qui retourne les listes de : du nombre de citations par année et les années de carrière de la personne
def donnees_citations_graph_citations(au_retrieval: AuthorRetrieval, document_eids: list):
    # Constantes nécessaires pour la suite des calculs
    first_year = au_retrieval.publication_range[0]
    total_annees = datetime.now().year - first_year + 2

    # Modifier les eid dans document_eids
    document_eids = [eid.split(".0-")[-1] for eid in document_eids]

    # Autres constantes nécessaires pour la suite des calculs
    length_list_eids = len(document_eids)
    stop_value = int((length_list_eids-1)/25)

    # Extraire les données du premier élément obligatoirement à part sinon cela impacte la boucle for si length > 25!
    co = CitationOverview(identifier=document_eids[0:1], start=first_year, end=first_year+total_annees-1, refresh=True)
    header_citation = co._header
    citation_overviews = []
    citation_overviews.append(co.cc)

    # Plusieurs extractions nécessaires si documents > 25
    if length_list_eids > 25:
        # Boucle le nombre de fois où il y a 25 dans length_list_eids
        for i in range(0, int(length_list_eids / 25)):
            # Extraire par 25 les données et les ajouter à la liste principale
            co = CitationOverview(identifier=document_eids[(i*25)+1 : (i*25)+26], start=first_year, end=first_year+total_annees-1, refresh=True)
            header_citation = co._header
            citation_overviews.append(co.cc)

    # Extraire les données qu'il reste (nb de documents < 25)
    if (length_list_eids-1)%25 != 0:
        co = CitationOverview(identifier=document_eids[stop_value*25 + 1 : stop_value*25 + (length_list_eids-1)%25 + 1], start=first_year, end=first_year+total_annees-1, refresh=True)
        header_citation = co._header
        citation_overviews.append(co.cc)
    else:
        citation_overviews.append([[(0, 0) for _ in range(total_annees)]])

    # Décapsulation de citation_overviews (liste principale) pour faire le total par année
    nb_cit_annees = [0] * total_annees
    for i in range(0, stop_value + 2):
        for y in range(0, len(citation_overviews[i])):
            for z in range(0, len(citation_overviews[i][y])):
                nb_cit_annees[z] += citation_overviews[i][y][z][1]

    # Affichage de manière tabulaire le nb de citations par année (T pour transposition matricielle)
    df = pd.DataFrame(nb_cit_annees).T # Transposition
    years_list = [first_year + i for i in range(0, total_annees)] # Modification des index de colonnes avec les années
    df.columns = years_list

    # Renommer l'axe des colonnes transposées ainsi que le nom de la colonne de données
    df = df.rename_axis('Année', axis='columns')
    df = df.rename(index={0: 'Citations'})

    # Ajoute le total de cette liste à la fin de la liste (écrasement/overwriting)
    nb_cit_annees.append(sum(nb_cit_annees))

    return nb_cit_annees, years_list, header_citation

# Fonction qui retourne le tableau pour le graphique des citations
def tab_graph_citations(au_retrieval: AuthorRetrieval, eids_list: list, liste_docs: list, console: QPlainTextEdit, window_width: int):
    # PARTIE sur les citations
    liste_citations, years_list, header = donnees_citations_graph_citations(au_retrieval, eids_list)

    # Créer une liste de paires avec les éléments alignés
    resultat = list(zip(liste_citations, liste_docs))

    # Affichage de manière tabulaire le nb de citations par année (T pour transposition matricielle)
    df = pd.DataFrame(resultat).T # Transposition
    years_list.append('TOTAL') # Ajout du nom de colonne TOTAL à la suite des années
    df.columns = years_list 

    # Renommer l'axe des colonnes transposées ainsi que le nom de la colonne de données
    df = df.rename_axis('Année', axis='columns')
    df = df.rename(index={0: 'Citations'})
    df = df.rename(index={1: 'Documents'})
    
    # Affichage
    console.append("\n" + '<p style="text-decoration: underline; color: black;">Tableau pour le graphique des <b>Citations</b> :</p>')
    console.append(df.to_string(index=True, col_space=0, line_width=window_width))

    return df, [au_retrieval.given_name, au_retrieval.surname], header


# Fonction utilitaire qui permet d'une liste de retourner une liste avec des 0
# à la place des éléments vides (NONE)
def _replace_none_with_zero(lst: list):
    for i in range(len(lst)):
        if lst[i] is None:
            lst[i] = 0
    return lst

# Fonction qui retourne les valeurs de l'encadré du rapport en fonction de l'eid de la personne sélectionnée
def valeurs_encadre(author_eid, years_list: list):
    # Instance de l'objet AuthorLookup correspondant à la personne sélectionnée via l'EID
    au = AuthorLookup(author_id=author_eid, refresh=True)

    # Obtient via l'instance les metrics "ScholarlyOutput" sur les 10 dernières années complètes sous forme de liste tot_scholarly_out
    liste_sch_out = _replace_none_with_zero(au.get_metrics_Other(metricType='ScholarlyOutput', yearRange='10yrs').List)

    # Adapte l'index correspondant à la première année pour les valeurs de l'encadré : moy de citations par pub ET moy MCR, vis-à-vis de la contrainte de l'API SciVal
    index_10y_adapted = liste_sch_out[0].index(years_list[0]) if years_list[0] in liste_sch_out[0] else 0
    # Créé une nouvelle liste en fonction de l'index trouvé
    annee_10y_adapt = liste_sch_out[0][index_10y_adapted]

    ### Publications très citées (1er décile) sur les 5 dernières années complètes - 1 ###
    # Calcul le total des "ScholarlyOutputs" pour les 5 dernières années complètes - 1
    tot_scholarly_5y = sum(liste_sch_out[-1][4:9])
    # Calcul le total des "OutputsInTopCitationPercentiles" pour les 5 dernières années complètes - 1
    tot_top_citations = sum(_replace_none_with_zero(au.get_metrics_Percentile(metricType='OutputsInTopCitationPercentiles', yearRange='10yrs').List[-1][4:9]))
    # Calcul final : rapport des totaux multiplié par 100 pour avoir en pourcentage et arrondi au dixième ET valeur mis à 0 si tot_scholarly_5y vaut 0
    top_citations = round(tot_top_citations/tot_scholarly_5y*100, 1) if tot_scholarly_5y != 0 else 0

    ### Publications en collaboration avec l'industrie de -6 ans à l'année prochaine ###
    # Calcul le total des "ScholarlyOutputs" pour les 5 dernières années - 1 complètes ainsi que l'année en cours et l'année future
    tot_scholarly_5ycf = sum([liste_sch_out[-1][4]] + _replace_none_with_zero(au.get_metrics_Other(metricType='ScholarlyOutput', yearRange='5yrsAndCurrentAndFuture').List[-1]))
    # Calcul le total des "AcademicCorporateCollaborations" pour les 5 dernières années - 1 complètes ainsi que l'année en cours et l'année future
    tot_acad_collab = sum(_replace_none_with_zero([au.get_metrics_Collaboration(metricType='AcademicCorporateCollaboration', yearRange='10yrs', collabType='Academic-corporate collaboration').List[-1][4]] + au.get_metrics_Collaboration(metricType='AcademicCorporateCollaboration', yearRange='5yrsAndCurrentAndFuture', collabType='Academic-corporate collaboration').List[-1]))
    # Calcul final : rapport des totaux multiplié par 100 pour avoir en pourcentage et arrondi au dixième ET valeur mis à 0 si tot_scholarly_5ycf vaut 0
    acad_collab = round(tot_acad_collab/tot_scholarly_5ycf*100, 1) if tot_scholarly_5ycf != 0 else 0

    ###### Moyennes de citations par publication ET MCR
    # Créé la liste des "ScholarlyOutputs" avec seulement les types Articles et ConferencePapers, et dynamiquement via l'index adapté
    liste_sch_out_10y_adapted_ArticlesConf = _replace_none_with_zero(au.get_metrics_Other(metricType='ScholarlyOutput', yearRange='10yrs', includedDocs='ArticlesConferencePapers').List)[-1][index_10y_adapted:9]
    # Constante résultante de la somme des éléments de la liste ci-dessus
    tot_liste_sch_out_10y_adapted_ArticlesConf = sum(liste_sch_out_10y_adapted_ArticlesConf)

    ### Moyenne de citations par publication ###
    # Créé la liste des "CitationsPerPublication" avec seulement les types Articles et ConferencePapers, et dynamiquement via l'index adapté
    liste_cit_per_pub = _replace_none_with_zero(au.get_metrics_Other(metricType='CitationsPerPublication', yearRange='10yrs', includedDocs='ArticlesConferencePapers').List[-1][index_10y_adapted:9])
    # Calcul final : produit du nombre de documents publiés par la moy de citations par pub. par année, divisé par le total de documents publiés sur le range d'années,
    # arrondi au dixième ET valeur mis à 0 si tot_liste_sch_out_10y_adapted_ArticlesConf vaut 0
    cit_per_pub = round(sum([elem1 * elem2 for elem1, elem2 in zip(liste_sch_out_10y_adapted_ArticlesConf, liste_cit_per_pub)])/tot_liste_sch_out_10y_adapted_ArticlesConf, 1) if tot_liste_sch_out_10y_adapted_ArticlesConf != 0 else 0

    ### Moyenne MCR ###
    # Créé la liste des "FieldWeightedCitationImpact" avec seulement les types Articles et ConferencePapers, et dynamiquement via l'index adapté
    liste_moy_MCR = _replace_none_with_zero(au.get_metrics_Other(metricType='FieldWeightedCitationImpact', yearRange='10yrs', includedDocs='ArticlesConferencePapers').List[-1][index_10y_adapted:9])
    # Calcul final : produit du nombre de documents publiés par la moy MCR par année, divisé par le total de documents publiés sur le range d'années,
    # arrondi au centième ET valeur mis à 0 si tot_liste_sch_out_10y_adapted_ArticlesConf vaut 0
    moy_MCR = round(sum([elem1 * elem2 for elem1, elem2 in zip(liste_sch_out_10y_adapted_ArticlesConf, liste_moy_MCR)])/tot_liste_sch_out_10y_adapted_ArticlesConf, 2) if tot_liste_sch_out_10y_adapted_ArticlesConf != 0 else 0

    return [top_citations, cit_per_pub, moy_MCR, acad_collab], annee_10y_adapt, au._header


# Fonction utilitaire pour gérer l'affichage mais aussi la création la liste des années sélectionnées
def _affichage_plages_annees(parts: list, selected_types: list, df: pd.DataFrame, console: QPlainTextEdit):
    # Constante nécessaire pour l'affichage et les calculs
    current_year = datetime.now().year

    # Affichage
    console.append('')
    console.append('<p><a style="font-weight: bold;">Votre sélection :</a> {}ans ({}), {}ans ({}) et Carrière ({})</p>'.format(current_year - parts[2], parts[2], current_year - parts[1], parts[1], parts[0]))
    console.append('\n')
    console.append('<p style="text-decoration: underline; color: black;">Nb de documents en fonction de leur type :</p>')

    # Filtrer le DataFrame en utilisant la méthode isin() avec la liste des index
    df_filtre_reset = df[df.index.isin(selected_types)].reset_index(drop=True)
    df_filtre_reset.index = range(len(df_filtre_reset))
    df_filtre_reset.index.name = 'Index'
    console.append(df_filtre_reset.to_string(index=True, col_space=0, line_width=200))

    # Créé la liste de listes comportant les plages d'années souhaitées pour la personne
    year_list = []
    for i in range(len(parts)):
        year_range = list(range(parts[i], current_year + 2))
        year_range.sort(reverse=True)
        year_list.append(year_range)

    return year_list, df_filtre_reset

# Fonction qui permet de retourner un booléen pour connaitre la validité de la commande de l'utilisateur,
# une liste de listes des plages d'années sélectionnées et un DataFrame avec les types de docs sélectionnés
def selection_plages_annees(annees_selec: str, years: list, selected_types: list, df: pd.DataFrame, console: QPlainTextEdit):
    # Séparer les types de documents sélectionnés par l'utilisateur (et supprimer les espaces avant et après les éléments)
    parts = annees_selec.split(',')
    parts = [element.strip() for element in parts]

    # Choix par défaut
    if len(parts) == 1 and parts[0] == "":
        return True, *_affichage_plages_annees([years[0], years[-1] - 6, years[-1] - 4], selected_types, df, console) # * permet d'ouvrir le tuple généré par la fonction
    
    # Si sélection avec manque ou surplus d'éléments
    if not(len(parts) == 2 or len(parts) == 3):
        # Ici indique erreur: manque ou surplus d'éléments
        console.append("<p style={}>! Manque ou surplus d'éléments (2 ou 3 éléments demandés)</p>".format(text_style_warning))
        return False, parts, df
    
    # Vérification de la validité des éléments
    tout_valide = True
    for year_index in parts:
        # Vérifier si l'index est valide
        if not (year_index.isdigit() and years[0] <= int(year_index) <= years[-1]):
            console.append('<p style={}>! Index non valide: {}</p>'.format(text_style_warning, year_index))
            tout_valide = False
    # Si au moins un élément n'est pas valide alors la sélection n'est pas retenue        
    if not tout_valide:
        return False, parts, df

    # Création de la liste utilisée par la suite pour filtrer les données voulues
    parts = [int(n) for n in parts]
    parts.append(years[0]) if len(parts) == 2 else None # Ajout année de début de carrière si 2 éléments
    parts.sort() # Trie de la liste pour s'assurer que les éléments soient dans le bon ordre

    return True, *_affichage_plages_annees(parts, selected_types, df, console) # * permet d'ouvrir le tuple généré par la fonction


# Fonction utilitaire pour créer une liste de listes en fonction des combinaisons sélectionnées
def _combine_types(chaine: str):
    main_indices_list = []
    # Création de la liste des types sélectionnés avec mise en forme (suppression des espaces)
    selected_types = chaine.split(',')
    selected_types = [element.strip() for element in selected_types]
    
    # Pour chaque type dans les types sélectionnés, vérifier le pattern
    for types in selected_types:
        pattern = r'\[(.*?)\]'
        contenu_crochets = re.findall(pattern, types)

        # Si pattern.s trouvé.s alors créer des listes dans la liste principale (Mise en forme relative à mon choix)
        if len(contenu_crochets) > 0:
            element_list = [element_of_element.strip() for element_of_element in contenu_crochets[0].split(';')]

            main_indices_list.append(element_list)
        else:
            main_indices_list.append([types])
    return main_indices_list

# Fonction qui retourne un booléen qui confirme la validité de la commande de l'utilisateur
# et les types de docs sélectionnés pour être mis en avant (combinaisons comprises)
def selection_2_types_docs(index_took: str, df: pd.DataFrame, console: QPlainTextEdit):
    # Combiner des types si c'est indiqué par l'utilisateur
    selected_types = _combine_types(index_took)
    
    tout_valide = True
    # Créé une liste des types de documents sélectionnés
    liste_types_selec = df['Type de documents'].to_list()

    # Si l'utilisateur prend les choix par défaut
    if len(selected_types) == 1 and selected_types[0][0] == "":
        console.append('')
        console.append('<p><a style="font-weight: bold;">Votre sélection :</a> {}, {}</p>'.format(liste_types_selec[0], liste_types_selec[1] if len(liste_types_selec)>1 else '∅'))
        console.append('\n')
        return True, [[liste_types_selec[0]], [liste_types_selec[1] if len(liste_types_selec)>1 else '∅']]
    
    # Si sélection avec manque ou surplus d'éléments
    if not len(selected_types) == 2:
        console.append("<p style={}>! Manque ou surplus d'éléments (2 éléments demandés)</p>".format(text_style_warning))
        return False, selected_types

    # Créé une liste "plate", "écrasée" de la liste de listes pour analyser élément par élément plus simplement
    flattened_list = [element for sublist in selected_types for element in sublist]
    for type_index in flattened_list:
        # Vérifier si l'index est valide
        if not (type_index.isdigit() and int(type_index) < len(df)):
            console.append('<p style={}>! Index non valide: {}</p>'.format(text_style_warning, type_index))
            tout_valide = False
            continue
        elif flattened_list.count(str(int(type_index))) > 1:
            console.append('<p style={}>! Doublon trouvé: {}</p>'.format(text_style_warning, type_index))
            tout_valide = False

    # Si au moins un élément n'est pas valide alors la sélection n'est pas retenue  
    if not tout_valide:
        return False, selected_types
    
    # Obtenir une liste d'entier puis mettre à jour la liste
    selected_types = [[int(element) for element in sublist] for sublist in selected_types]

    # Affichage
    console.append('')
    console.append('<p><a style="font-weight: bold;">Votre sélection :</a> {}, {}</p>'.format(df.at[selected_types[0][0], 'Type de documents'], df.at[selected_types[1][0], 'Type de documents']))
    console.append('\n')

    return True, [[df.loc[element, 'Type de documents'] for element in sublist] for sublist in selected_types] # Traduction des index en nom de type de documents




# Fonction qui retourne le tableau pour le graphique des publications
def tab_graph_publications(au_retrieval: AuthorRetrieval, document_eids: list, liste_annees: list, liste_type: list, console: QPlainTextEdit, window_width: int):
    # Parcourir chaque sous-liste de la liste pour modifier les les types des années (de int à str)
    liste_annees = [[str(annee) for annee in sous_liste] for sous_liste in liste_annees]

    # Filtrer les documents en fonction des EIDs spécifiés
    docs = pd.DataFrame(au_retrieval.get_documents(refresh=10))
    docs_filtered = docs[docs['eid'].isin(document_eids)]

    # Extraire les années de publication
    df = pd.DataFrame({
        'DocType': docs_filtered['subtypeDescription'],
        'Year': pd.to_datetime(docs_filtered['coverDate']).dt.year.astype(str),
    })
    # Traduction inverse pour les matchs après
    liste_type_en = [[trad_fr2en[doc] for doc in sublist] for sublist in liste_type]
    # Créé une liste "plate", "écrasée" de la liste de listes pour plus simplement manier les données des types
    flattened_liste_type_en = [element for sublist in liste_type_en for element in sublist]

    # DataFrame vide pour stocker les résultats
    results = pd.DataFrame(columns=flattened_liste_type_en + ['Autres'])  

    # Pour chaque range d'années sélectionné, faire les totaux des documents par type (de tous les index demandés)
    for annees in liste_annees:
        df_filtered = df[df['Year'].isin(annees)]
        counts = df_filtered['DocType'].value_counts()
        total_doctype = len(df_filtered)

        liste = counts.reindex(flattened_liste_type_en, fill_value=0).tolist()
        liste.append(total_doctype - sum(liste))

        df_result = pd.DataFrame([liste], columns=flattened_liste_type_en + ['Autres'])
        results = pd.concat([results, df_result], ignore_index=True)
        
    # S'il y a des combinaisons de types alors on additionne les colonnes sous le nom du premier et on supprime les colonnes qui se combinent au premier
    for sublist in liste_type_en:
        for i in range(1, len(sublist)):
            results[sublist[0]] += results[sublist[i]]
            results = results.drop(sublist[i], axis=1)

    # Renommage dynamique des index et des colonnes du dataframe
    results = results.rename(index={0: 'Carrière (≥' + liste_annees[0][-1] + ')', 1: liste_annees[1][-1] + ' à ≥' + liste_annees[0][1], 2: liste_annees[2][-1] + ' à ≥' + liste_annees[0][1]}, 
                             columns={liste_type_en[0][0]: liste_type[0][0], liste_type_en[1][0]: liste_type[1][0]})

    # Calcul le total par ligne et ajouter une colonne 'Total'
    results = results.assign(TOTAL=results.sum(axis=1))

    # Affiche le tableau
    console.append("\n" + '<p style="text-decoration: underline; color: black;">Tableau pour le graphique des <b>Publications</b> :</p>')
    console.append(results.to_string(index=True, col_space=0, line_width=window_width))

    return results


# Fonction utilitaire de la fonction "tab_graph_SNIP" pour permettre d'extraire depuis un résultat
# d'une requête les valeurs nécessaires pour les calculs pour le graphique SNIP
def _for_SNIP_list_10y_current_future(lst: list):
    # Pour chaque élément de la liste (qui sont des dictionnaires)
    for element in lst:
        # Prend les valeurs par années
        value_by_year = element['valueByYear']

        if element['threshold'] == 5:
            annees = [int(annee) for annee in list(value_by_year.keys())]   # Créé la liste du range d'années
            element_with_threshold_5 = list(value_by_year.values())         # Créé la liste des valeurs pour le seuil des 5%
        elif  element['threshold'] == 10:
            element_with_threshold_10 = list(value_by_year.values())        # Créé la liste des valeurs pour le seuil des 10%
        elif  element['threshold'] == 25:
            element_with_threshold_25 = list(value_by_year.values())        # Créé la liste des valeurs pour le seuil des 25%

    return [annees, element_with_threshold_5, element_with_threshold_10, element_with_threshold_25]

# Fonction qui retourne un DataFrame (tableau) pour le graphique SNIP du rapport
def tab_graph_SNIP(author_id: str, years_list: list, console: pd.DataFrame, window_width: int):
    # Convertie le type toutes les années (de str/string à int/integer)
    years_list = [[int(item) for item in sublist] for sublist in years_list]

    # Instance de l'objet AuthorLookup correspondant à la personne sélectionnée via l'ID
    au = AuthorLookup(author_id=author_id, refresh=True)

    # Obtient via l'instance les metrics "PublicationsInTopJournalPercentiles" avec seulement les types Articles et Reviews
    # sur les 10 dernières années complètes sous forme de liste ten_y_cf_list
    ten_y_cf_list = _for_SNIP_list_10y_current_future(au._get_metrics_rawdata(metricType='PublicationsInTopJournalPercentiles', yearRange='10yrs', journalImpactType="SNIP", includedDocs='ArticlesReviews'))
    # Même chose pour 3 years and current and future
    three_y_cf_list = _for_SNIP_list_10y_current_future(au._get_metrics_rawdata(metricType="PublicationsInTopJournalPercentiles", yearRange="3yrsAndCurrentAndFuture", journalImpactType="SNIP", includedDocs='ArticlesReviews'))

    # Obtient via l'instance les metrics "ScholarlyOutput" avec seulement les types Articles et Reviews
    # sur les 10 dernières années complètes sous forme de liste ten_y_scho_list
    ten_y_scho_list = au.get_metrics_Other(metricType="ScholarlyOutput", yearRange="10yrs", includedDocs='ArticlesReviews').List
    # Même chose pour 3 years and current and future
    three_y_cf_scho_list = au.get_metrics_Other(metricType="ScholarlyOutput", yearRange="3yrsAndCurrentAndFuture", includedDocs='ArticlesReviews').List

    # Concaténation des listes en une seule sous le format : [[années], [ScholarlyOutputs], [Top5%], [Top10%], [Top25%]]
    ten_y_cf_list = [_replace_none_with_zero(item1 + item2[-2:]) for item1, item2 in zip(ten_y_scho_list, three_y_cf_scho_list)] + [_replace_none_with_zero(item1 + item2[-2:]) for item1, item2 in zip(ten_y_cf_list, three_y_cf_list)][-3:]

    # Index par défaut si les index choisis par l'utilisateur ne rentrent pas dans le range max de 10y, de la contrainte des API SciVal
    default_index_year_list = [0, -7, -5]
    # Réalisation de la liste des années sélectionnées si c'est dans le range max proposé par les API SciVal (ex: [2013, 2018, 2020])
    real_years_list = [years_list[i][-1] if years_list[i][-1] in ten_y_cf_list[0] else ten_y_cf_list[0][default_index_year_list[i]] for i in range(3)]
    
    # Liste des noms des colonnes
    column_name_list = ['Top 5%', 'Top 10%', 'Top 25%', 'Autres']
    # Créé le DataFrame avec toutes nos données récoltées
    df = pd.DataFrame().assign(**{column_name_list[i]: [sum(ten_y_cf_list[i+2][ten_y_cf_list[0].index(year):]) for year in real_years_list] for i in range(len(column_name_list) - 1)})

    # Créé la colonne Autres et réalise les vraies colonnes de données Top 25% et Top 10%
    df[column_name_list[3]] = [sum(ten_y_cf_list[1][ten_y_cf_list[0].index(year):]) - df['Top 25%'].to_list()[j] for j, year in enumerate(real_years_list)]
    df['Top 25%'] = df['Top 25%'] - df['Top 10%']
    df['Top 10%'] = df['Top 10%'] - df['Top 5%']

    # Renomme les index de manière dynamique
    df = df.rename(index={0: '≥'+str(real_years_list[0]), 1: str(real_years_list[1])+' à ≥'+str(ten_y_cf_list[0][-2]), 2: str(real_years_list[2])+' à ≥'+str(ten_y_cf_list[0][-2])})

    # Ajoute une colonne "TOTAL" contenant la somme des valeurs des autres colonnes
    df['TOTAL'] = df.sum(axis=1)

    # Affiche le tableau
    console.append('\n')
    console.append('<p style="text-decoration: underline; color: black;">Tableau pour le graphique <b>SNIP</b> :</p>')
    console.append(df.to_string(index=True, col_space=0, line_width=window_width))

    return df, au._header


# Fonction utilitaire de la fonction "tab_graph_Collab" pour permettre d'extraire depuis un résultat
# d'une requête les valeurs nécessaires pour les calculs pour le graphique SNIP
def _for_Collab_list_10y_current_future(lst: list):
    # Pour chaque élément de la liste (qui sont des dictionnaires)
    for element in lst:
        # Prend les valeurs par années
        value_by_year = element['valueByYear']

        if element['collabType'] == "Institutional collaboration":
            annees = [int(annee) for annee in list(value_by_year.keys())]   # Créé la liste du range d'années
            inst_collab = list(value_by_year.values())                      # Créé la liste des valeurs pour la collaboration institutionnelle
        elif  element['collabType'] == "International collaboration":
            international_collab = list(value_by_year.values())             # Créé la liste des valeurs pour la collaboration internationale
        elif  element['collabType'] == "National collaboration":
            national_collab = list(value_by_year.values())                  # Créé la liste des valeurs pour la collaboration nationale
        elif  element['collabType'] == "Single authorship":
            no_collab = list(value_by_year.values())                        # Créé la liste des valeurs pour les publications sans collaboration

    return [annees, inst_collab, international_collab, national_collab, no_collab]

# Fonction qui retourne un DataFrame (tableau) pour le graphique Collaborations du rapport
def tab_graph_Collab(author_id: str, years_list: list, console: pd.DataFrame, window_width: int):
    # Instance de l'objet AuthorLookup correspondant à la personne sélectionnée via l'ID
    au = AuthorLookup(author_id=author_id, refresh=True)

    # Obtient via l'instance les metrics "Collaboration" sur les 10 dernières années complètes sous forme de liste ten_y_cf_list
    ten_y_cf_list = _for_Collab_list_10y_current_future(au._get_metrics_rawdata(metricType='Collaboration', yearRange='10yrs'))
    # Même chose pour 3 years and current and future
    three_y_cf_list = _for_Collab_list_10y_current_future(au._get_metrics_rawdata(metricType='Collaboration', yearRange='3yrsAndCurrentAndFuture'))

    # Concaténation des listes en une seule sous le format : [[années], [Inst], [Inter], [Nat], [Aucune]]
    ten_y_cf_list = [_replace_none_with_zero(item1 + item2[-2:]) for item1, item2 in zip(ten_y_cf_list, three_y_cf_list)]

    # Index par défaut si les index choisis par l'utilisateur ne rentrent pas dans le range max de 10y, de la contrainte des API SciVal
    default_index_year_list = [0, -7, -5]
    # Réalisation de la liste des années sélectionnées si c'est dans le range max proposé par les API SciVal (ex: [2013, 2018, 2020])
    real_years_list = [years_list[i][-1] if years_list[i][-1] in ten_y_cf_list[0] else ten_y_cf_list[0][default_index_year_list[i]] for i in range(3)]

    # Liste des noms des colonnes
    column_name_list = ['Internat.', 'Nationale', 'Inst.', 'Aucune']
    # Créé le DataFrame avec toutes nos données récoltées
    df = pd.DataFrame().assign(**{column_name_list[i]: [sum(ten_y_cf_list[i+1][ten_y_cf_list[0].index(year):]) for year in real_years_list] for i in range(len(column_name_list))})

    # Intervertie les colonnes pour avoir l'ordre: International, National, Institutionnel
    df['Internat.'], df['Nationale'], df['Inst.'] = df['Nationale'].copy(), df['Inst.'].copy(), df['Internat.'].copy()

    # Renomme les index
    df = df.rename(index={0: '≥'+str(real_years_list[0]), 1: str(real_years_list[1])+' à ≥'+str(ten_y_cf_list[0][-2]), 2: str(real_years_list[2])+' à ≥'+str(ten_y_cf_list[0][-2])})

    # Ajoute une colonne "Total" contenant la somme des valeurs des autres colonnes
    df['TOTAL'] = df.sum(axis=1)

    # Affiche le tableau
    console.append('\n')
    console.append('<p style="text-decoration: underline; color: black;">Tableau pour le graphique des <b>Collaborations</b> :</p>')
    console.append(df.to_string(index=True, col_space=0, line_width=window_width))

    return df, au._header




# Fonction qui permet d'exporter les données sur le gabarit Excel et d'appeler les
# routines VBA du gabarit
def Excel_part1(df: pd.DataFrame, nom_prenom: list, en_tete: list, annee_10y_adapt: int):
    # Ouvrir le classeur Excel existant
    nom_fichier = os.path.dirname(os.path.abspath(__file__)) + '\\..\\GABARIT.xlsm'
    nom_feuille = 'Raw_Data'
    nom_module = 'Module1'
    nom_procedure = 'AjusterDynamiquementAbscisseGraphiqueCitations'

    cell_tab_citations = [4, 0] # ligne, colonne

    # Création de l'objet Excel, et le rendre visible en plein écran lors du processus
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = True
    excel.WindowState = win32.constants.xlMaximized

    try:
        # Vérifier si le fichier Excel est déjà ouvert
        for wb in excel.Workbooks:
            if wb.FullName == nom_fichier:
                wb.Close(False)  # Fermer le classeur sans enregistrer les modifications

        # Ouverture du fichier Excel
        classeur = excel.Workbooks.Open(nom_fichier)
        classeur.Visible = True  # Rendre le classeur visible
        classeur.WindowState = win32.constants.xlMaximized  # Mettre le classeur en plein écran 

        # Enregistrer sous le nouveau nom
        date_formated = datetime.now().strftime('%Y-%m-%d')
        nom_ou_prenom = []
        for i in range(len(nom_prenom)):
            # Supprimer tous les accents, ...
            nom_ou_prenom.append(unicodedata.normalize("NFD", nom_prenom[i]).encode("ascii", "ignore").decode("utf-8")) # une str sans accent
            # Remplacer les espaces par des tirets
            nom_ou_prenom[i] = nom_ou_prenom[i].replace(' ', '-')
            # Supprimer tous les apostrophes
            nom_ou_prenom[i] = nom_ou_prenom[i].replace("'", '')

        classeur.SaveAs(os.path.abspath(DOCS_PATH[0] + '/' + date_formated + '_' + nom_ou_prenom[1] + '_' + nom_ou_prenom[0] + '.xlsm'), FileFormat=52)

        # Mettre la fenêtre en premier plan
        try:
            win32gui.SetForegroundWindow(win32gui.FindWindow(None, classeur.Name + " - Excel"))
        except:
            win32gui.SetForegroundWindow(win32gui.FindWindow(None, classeur.Name.split('.')[0] + " - Excel"))
      
        # Accéder à la feuille de calcul existante
        feuille = classeur.Worksheets(nom_feuille)

        # Effacer le contenu des lignes de cellules et cellules à modifier
        start_row = 2 + cell_tab_citations[0]  # Numéro de la première ligne de cellule (à supprimer)
        end_row = start_row + len(df) - 1  # Numéro de la dernière ligne de cellule (à supprimer)

        feuille.Range(f"{start_row-1}:{end_row}").ClearContents()  # Nettoie seulement le contenu des lignes souhaitées

        # Path ainsi que le path de ce programme pour enregistrer le Word
        feuille.Cells(1, 110).Value = DOCS_PATH[0]
        feuille.Cells(2, 110).Value = os.path.dirname(os.path.abspath(__file__)) + "\\..\\"

        # Écrire prénom et nom
        for i in range(len(nom_prenom)):
            feuille.Cells(1 + i, 2).Value = nom_prenom[i]

        # Écrire année des 10y_adapt
        feuille.Cells(1, 104).Value = annee_10y_adapt

        # Écrire prénom et nom sans accent
        for i in range(len(nom_prenom)):
            feuille.Cells(1 + i, 5).Value = nom_ou_prenom[i]

        # Écrire les données du DataFrame dans la feuille de calcul
        for i, row in enumerate(df.values):
            for j, value in enumerate(row):
                if j > len(row) - 4:
                    feuille.Cells(i+2+cell_tab_citations[0], j+2+cell_tab_citations[1]+2).Value = value  # Ajouter 2 pour décaler les cellules
                    if j > len(row) - 3:
                        continue
                feuille.Cells(i+2+cell_tab_citations[0], j+2+cell_tab_citations[1]).Value = value

        # Écrire les index dans la première colonne de la feuille de calcul
        for i, index_value in enumerate(df.index):
            feuille.Cells(i+2+cell_tab_citations[0], 1+cell_tab_citations[1]).Value = index_value

        # Écrire les données supplémentaires
        year_list = df.columns.to_list()
        feuille.Cells(1, 103).Value = year_list[-3]
        feuille.Cells(2, 102).Value = year_list[0]
        feuille.Cells(1, 106).Value = date_formated
        # Convertir la chaîne de date en objet datetime
        date_obj = datetime.strptime(date_formated, '%Y-%m-%d')
        feuille.Cells(1, 107).Value = date_obj.day
        feuille.Cells(1, 108).Value = date_obj.month
        feuille.Cells(1, 109).Value = date_obj.year


        # Écrire les noms des colonnes dans la première ligne de la feuille de calcul d'après les conditions du cahier des charges
        len_row = len(row)
        len_year_list = len(year_list)

        # Écrit TOUS les noms des colonnes
        for j, column_name in enumerate(df.columns):
            column_name = str(column_name)[-2:] if len_year_list-3 <= 20 and j < len_row-1 else column_name

            if j > len_row-4:
                feuille.Cells(1+cell_tab_citations[0], j+2+cell_tab_citations[1] +2).Value = column_name
                if j > len_row-3:
                    continue
            feuille.Cells(1+cell_tab_citations[0], j+2+cell_tab_citations[1]).Value = column_name

        # Efface si nécessaire en fonction du CDC
        if len_year_list-3 > 10:
            for j, column_name in enumerate(df.columns):
                if len_year_list-3 <= 30 and (j%2 == 1 and j != len_year_list-3 or (len_year_list-2)%2 == 0 and j == len_year_list-4) or len_year_list-3 > 30 and (j%5 != 0 and j != len_year_list-3 or (len_year_list-2)%2 == 0 and j == len_year_list-5 or len_year_list-7 < j < len_year_list-3):
                    feuille.Cells(1+cell_tab_citations[0], j+2+cell_tab_citations[1]).Value = None


        # Écrire les données de l'en-tête de SciVal
        for i in range(len(en_tete)):
            feuille.Cells(32 + i, 2).Value = en_tete[i]

        # Appel de la procédure VBA
        excel.Run(f'{nom_module}.{nom_procedure}', nom_feuille, 'TOTAL', 5)
        
        classeur.Visible = False  # Rendre le classeur invisible
        excel.Visible = False

    except Exception as e:
        print(f"Une erreur s'est produite : {e}")

    return excel, classeur

    
# Fonction qui reprend le classeur ouvert (caché) et qui permet d'exporter le reste des 
# données sur le gabarit Excel et d'appeler la routine VBA du gabarit Excel qui
# remplie le gabarit Word pour avoir la fiche bibliométrique finale!
def Excel_part2(excel, classeur, df: pd.DataFrame, df_SNIP: pd.DataFrame, df_Collab: pd.DataFrame):
    # Ouvrir le classeur Excel existant
    nom_fichier = os.path.dirname(os.path.abspath(__file__)) + '\\..\\GABARIT.xlsm'
    nom_feuille = 'Raw_Data'
    nom_module = 'Module1'

    cell_tab_publications = [10, 41, 48] # lignes de commencement des tableaux pour les graphiques

    # Rendre à nouveau visible l'Excel
    excel.Visible = True

    try:
        # Vérifier si le fichier Excel est déjà ouvert
        for wb in excel.Workbooks:
            if wb.FullName == nom_fichier:
                wb.Close(False)  # Fermer le classeur sans enregistrer les modifications

        # Ouverture du fichier Excel
        classeur.Visible = True  # Rendre le classeur visible
        classeur.WindowState = win32.constants.xlMaximized  # Mettre le classeur en plein écran 

        # Mettre la fenêtre Excel en premier plan
        try:
            win32gui.SetForegroundWindow(win32gui.FindWindow(None, classeur.Name + " - Excel"))
        except:
            win32gui.SetForegroundWindow(win32gui.FindWindow(None, classeur.Name.split('.')[0] + " - Excel"))


        # Accéder à la feuille de calcul existante
        feuille = classeur.Worksheets(nom_feuille)


        ### GRAPHIQUE DES PUBLICATIONS ###

        # Effacer le contenu des lignes de cellules et cellules à modifier
        start_row = 2 + cell_tab_publications[0]  # Numéro de la première ligne de cellule (à supprimer)
        end_row = start_row + len(df) - 1  # Numéro de la dernière ligne de cellule (à supprimer)

        feuille.Range(f"{start_row-1}:{end_row}").ClearContents()   # Nettoie seulement le contenu des lignes souhaitées

        # Écrire les données du DataFrame dans la feuille de calcul
        for i, row in enumerate(df.values):
            for j, value in enumerate(row):
                if not value == 0:
                    feuille.Cells(i+2+cell_tab_publications[0], j+2).Value = value
                else:
                    feuille.Cells(i+2+cell_tab_publications[0], j+2).ClearContents()

        # Écrire les index dans la première colonne de la feuille de calcul
        for i, index_value in enumerate(df.index):
            feuille.Cells(i+2+cell_tab_publications[0], 1).Value = index_value

        # Écrire les noms des colonnes dans la première ligne de la feuille de calcul
        for j, column_name in enumerate(df.columns):
            feuille.Cells(1+cell_tab_publications[0], j+2).Value = column_name


        ### GRAPHIQUE SNIP ###

        # Effacer le contenu des lignes de cellules et cellules à modifier
        start_row = 2 + cell_tab_publications[1]  # Numéro de la première ligne de cellule (à supprimer)
        end_row = start_row + len(df_SNIP) - 1  # Numéro de la dernière ligne de cellule (à supprimer)

        feuille.Range(f"{start_row-1}:{end_row}").ClearContents()   # Nettoie seulement le contenu des lignes souhaitées

        # Écrire les données du DataFrame dans la feuille de calcul
        for i, row in enumerate(df_SNIP.values):
            for j, value in enumerate(row):
                if not value == 0:
                    feuille.Cells(i+2+cell_tab_publications[1], j+2).Value = value
                else:
                    feuille.Cells(i+2+cell_tab_publications[1], j+2).ClearContents()

        # Écrire les index dans la première colonne de la feuille de calcul
        for i, index_value in enumerate(df_SNIP.index):
            feuille.Cells(i+2+cell_tab_publications[1], 1).Value = index_value

        # Écrire les noms des colonnes dans la première ligne de la feuille de calcul
        for j, column_name in enumerate(df_SNIP.columns):
            feuille.Cells(1+cell_tab_publications[1], j+2).Value = column_name


        ### GRAPHIQUE COLLAB ###

        # Effacer le contenu des lignes de cellules et cellules à modifier
        start_row = 2 + cell_tab_publications[2]  # Numéro de la première ligne de cellule (à supprimer)
        end_row = start_row + len(df_Collab) - 1  # Numéro de la dernière ligne de cellule (à supprimer)

        feuille.Range(f"{start_row-1}:{end_row}").ClearContents()   # Nettoie seulement le contenu des lignes souhaitées

        # Écrire les données du DataFrame dans la feuille de calcul
        for i, row in enumerate(df_Collab.values):
            for j, value in enumerate(row):
                if not value == 0:
                    feuille.Cells(i+2+cell_tab_publications[2], j+2).Value = value
                else:
                    feuille.Cells(i+2+cell_tab_publications[2], j+2).ClearContents()

        # Écrire les index dans la première colonne de la feuille de calcul
        for i, index_value in enumerate(df_Collab.index):
            feuille.Cells(i+2+cell_tab_publications[2], 1).Value = index_value

        # Écrire les noms des colonnes dans la première ligne de la feuille de calcul
        for j, column_name in enumerate(df_Collab.columns):
            feuille.Cells(1+cell_tab_publications[2], j+2).Value = column_name

        # Ouvrir Word
        word_app = win32.Dispatch("Word.Application")

        # Fermer les documents
        for doc in word_app.Documents:
            doc.Close(SaveChanges=True)

        # Reboot pour prendre la main
        word_app.Quit()

        # Appel de la procédure VBA
        excel.Run(f'{nom_module}.GenerationWord')

        # Attendre la fin des subroutines en cours
        time.sleep(1)
        
        # Récupérer la feuille à supprimer
        sheet = wb.Sheets("Main")
        # Supprimer la feuille
        excel.DisplayAlerts = False  # Désactiver les alertes
        sheet.Delete()
        excel.DisplayAlerts = True  # Réactiver les alertes

        # Supprimer les cellules tampons
        start_row, num_rows = 1, 2
        start_col, num_cols = 100, 11

        for row in range(start_row, start_row + num_rows):
            for col in range(start_col, start_col + num_cols):
                feuille.Cells(row, col).ClearContents()

        # Enregistrer et fermer le classeur Excel
        classeur.Close(SaveChanges=True)

        # Fermeture de l'application Excel
        excel.Quit()

    except Exception as e:
        print(f"Une erreur s'est produite : {e}")


#-------------------------------------Nouvelles fonctions d'Autobib+-------------------------------------------------

def collaborationExtract(researchersA: list = None, institutionsA: list = None, researchersB: list = None, institutionsB: list = None,\
                         country: str = None, start_year: int = None, end_year: int = None, keys: list = None, console: QPlainTextEdit = None):
    # Construction de la requete pour les collabs entre l'entité A et l'entité B
    query_part2 = []
    if researchersA:
        query_part1 = [f'AU-ID({researcher})' for researcher in researchersA]
    elif institutionsA:
        query_part1 = [f'AF-ID({institution})' for institution in institutionsA]
    else :
        return
    query_partA = " OR ".join(query_part1)
    
    if country:
        query_part2.append(f'AFFILCOUNTRY({country})')
    
    elif researchersB:
        researcher_query = " OR ".join([f'AU-ID({researcher})' for researcher in researchersB])
        query_part2.append(f'({researcher_query})')
    
    elif institutionsB:
        institution_query = " OR ".join([f'AF-ID({institution})' for institution in institutionsB])
        query_part2.append(f'({institution_query})')
    
    if start_year:
        query_part2.append(f'PUBYEAR AFT {start_year-1}')
    
    if end_year:
        query_part2.append(f'PUBYEAR BEF {end_year+1}')

    query_partB = " AND ".join(query_part2)
    
    RequestQuery = f"({query_partA}) AND ({query_partB})"
    
    try:
        # Recherche sur Scopus avec la clé API et le Token
        search = ScopusSearch(query=RequestQuery, api_key= keys[0], token= keys[1])
        if search.results is not None:
            # Extraction des résultats
            results = []
            for collaboration in search.results:
                results.append({
                    'EID': collaboration.eid,
                    'Abstract': getAbstract(collaboration.eid, keys),
                    'Title': collaboration.title,
                    # 'Source title': collaboration.subtype,
                    'Authors': collaboration.author_names,
                    'Authors ID': collaboration.author_ids,
                    'Authors affiliations': collaboration.author_afids,
                    'PublicationName': collaboration.publicationName,
                    'Year': collaboration.coverDisplayDate[-4:],  # Garder les 4 derniers chiffres de la date (l'annee)
                    'Cited By': collaboration.citedby_count,
                    'DOI': collaboration.doi,
                    'Authors keywords': collaboration.authkeywords, 
                    # 'AggregationType': collaboration.aggregationType,
                    'Nbre de publications': collaboration.afid,
                    'affilname': collaboration.affilname,
                    'Countries': collaboration.affiliation_country,
                    'DocumentType': collaboration.subtype, 
                    'Funding details': collaboration.fund_sponsor,
                    'Funding texts': collaboration.fund_acr,
                })
            
            # Conversion des résultats en DataFrame pandas
            df = pd.DataFrame(results)
            return df
        else: 
            console.append('<p style={}>! Aucune collaboration trouvée.</p>'.format(text_style_warning))
            return None
    except Exception as e: 
        return None
# Retourne des informations sur le profil recherché 
# On peut faire la recherche à partir d'un nom d'un chercher ou d'un identifiant  
def getEntityProfile(selection: str, entity: str, keys: list, rechercheParId: bool):

    if selection == '1':
        if rechercheParId is False : 
            nomComplet = entity.split(',')
            query_entity = f'authlast({nomComplet[0]}) and authfirst ({nomComplet[1]})'
            search = AuthorSearch(query=query_entity, api_key= keys[0], token= keys[1], refresh=True)
        else : 
            query_entity = f'AU-ID({entity})'
            search = AuthorSearch(query=query_entity, api_key= keys[0], token= keys[1])
        # Vérification des résultats
        if search.authors is None:
            return 'NONE'
        else:
            author_info = []
            for element in search.authors:
                author_info.append(
                    f"Nom : {element.givenname} {element.surname}\n"
                    f"ID : {element.eid}\n"
                    f"ORCID : {element.orcid}\n"
                    f"Affiliation : {element.affiliation}\n"
                    f"Nombre total de documents : {element.documents}\n"
                )
            return author_info
    elif selection == '2':
        query_entity = f'AF-ID({entity})'
        search = AffiliationSearch(query=query_entity, api_key= keys[0], token= keys[1])
        if search.affiliations is None:
            return 'NONE'
        else:
            for element in search.affiliations:
                institution_info = (
                    f"Nom : {element.name}\n"
                    f"ID : {element.eid}\n"
                    f"Localisation : {element.city}, {element.country}\n"
                    # f"Parent : {element.parent}\n"
                )
            return institution_info
    else :
        return
# Limite les plages de colloborations à année courante -20 , année courante +1
def getSelectedYears(response:str):
        if response == '' :
            start_year = datetime.now().year-5
            end_year = datetime.now().year
        elif ',' in response: 
            selected_years = response.split(',')
            # Recuperer les limites de la plage séléctionnée
            start_year = int(selected_years[0])
            end_year = int(selected_years[1])
            if start_year < datetime.now().year-20 or end_year > datetime.now().year + 1: 
                start_year = 'NULL'
                end_year = 'NULL'
        else:
            start_year = 'NULL'
            end_year = 'NULL'
        return start_year, end_year

def count_document_types(df: pd.DataFrame):
    # Initialiser un dictionnaire pour stocker les comptes de chaque type de document
    doc_type_counts = {
        'ar': 0,
        're': 0,
        'cp': 0,
        'ch': 0,
        'ed': 0,
        'bk': 0,
        'dp': 0,
        'er': 0,
        'sh': 0, 
    }
    
    if 'DocumentType' in df.columns:
        # Compter les occurrences de chaque type de document
        doc_type_counts = df['DocumentType'].value_counts().to_dict()
        
        # Ajouter les types de documents manquants avec un compte de 0
        for doc_type in ['ar', 're', 'cp', 'ch', 'ed', 'bk', 'dp', 'er', 'sh']:
            if doc_type not in doc_type_counts:
                doc_type_counts[doc_type] = 0
        # Créer une DataFrame à partir du dictionnaire
        doc_type_df = pd.DataFrame(list(doc_type_counts.items()), columns=['DocumentType', 'Count'])
        
        # Ajouter cette nouvelle DataFrame à la DataFrame originale (comme une nouvelle colonne)
        df_with_counts = pd.concat([df, doc_type_df], axis=1)
        # Remplacer les NaN par une chaîne vide
        df_with_counts = df_with_counts.fillna('')

    # Retourner la DataFrame originale
    return df_with_counts


def countAuthorsInCollab(df : pd.DataFrame, keys: list):
        author_counts = {}
        Author_IDs = {}
        Author_Aff = {}
        if 'Authors' in df.columns:
            authors = df['Authors'].dropna().str.split(';')
            authorsIDs = df['Authors ID'].dropna().str.split(';')
            afIDs = df['Authors affiliations'].dropna().str.split(';')
            for afID_list, author_list, authorsID_list in zip(afIDs, authors, authorsIDs):
                for afID, author, authorID in zip(afID_list, author_list, authorsID_list):
                    author = author.strip()
                    authorID = authorID.strip()
                    afID = afID.strip()
                    if not afID == '':
                        name_parts = author.split(', ')
                        last_name = name_parts[0].strip()
                        first_name = name_parts[1].strip() if len(name_parts) > 1 else ""
                        update_entity_author_counts(author_counts, Author_IDs, 
                                                                Author_Aff, last_name, first_name, 
                                                                author, authorID, keys)
            author_df = pd.DataFrame(list(author_counts.items()), columns=['Author', 'Nbre de publications'])
             # Ajouter les colonne pour les IDs , affiliation des auteurs
            author_df['AU-ID'] = author_df['Author'].map(lambda author: ', '.join(Author_IDs.get(author, [])))
            author_df['AU-ID']  = author_df['AU-ID'].str.replace(', ', '', regex=False)
            author_df['Affiliation'] = author_df['Author'].map(lambda author: ', '.join(Author_Aff.get(author, [])))
            author_df['Affiliation']  = author_df['Affiliation'].str.replace(', ', '', regex=False)
            # Trier les résultats
            author_df = author_df.sort_values(by='Nbre de publications', ascending=False).reset_index(drop=True)
            return author_df
        
def countInstitutionsInCollab(df : pd.DataFrame, collabCountry : str):
        institution_counts = {}
        if 'affilname' in df.columns and 'Countries' in df.columns:
            affiliations = df['affilname'].dropna().str.split(';')
            countries = df['Countries'].dropna().str.split(';')
                
            for affil_list, country_list in zip(affiliations, countries):
                for affil, country in zip(affil_list, country_list):
                    affil = affil.strip()
                    if country.strip() == collabCountry:
                        if affil in institution_counts:
                            institution_counts[affil] += 1
                        else:
                            institution_counts[affil] = 1

            institution_df = pd.DataFrame(list(institution_counts.items()), columns=['Institution', 'Nbre de publications en collaboration'])
            institution_df = institution_df.sort_values(by='Nbre de publications en collaboration', ascending=False).reset_index(drop=True)
            return institution_df

def countEntityAuthorsInCollab(df : pd.DataFrame, collabEntityList : list, keys: list):
    entityAuthor_counts = {}
    entityAuthor_IDs = {}
    entityAuthor_Aff = {}
    if 'Authors' in df.columns and 'Authors affiliations' in df.columns:
        authorsIDs = df['Authors ID'].dropna().str.split(';')
        afIDs = df['Authors affiliations'].dropna().str.split(';')
        authors = df['Authors'].dropna().str.split(';')
        # eids = df['EID'].dropna().str.split(';')
        for collabEntity in collabEntityList:
            for afID_list, author_list, authorsID_list in zip(afIDs, authors, authorsIDs):
                for afID, author, authorID in zip(afID_list, author_list, authorsID_list):
                    author = author.strip()
                    authorID = authorID.strip()
                    afID = afID.strip()
                    # collabEntity = collabEntity
                    collabEntity = collabEntity.strip()
                    name_parts = author.split(', ')
                    last_name = name_parts[0].strip()
                    first_name = name_parts[1].strip() if len(name_parts) > 1 else ""
                    if '-' in afID:
                        affs = afID.split('-')
                        for aff in affs:
                            if aff == collabEntity:
                                update_entity_author_counts(entityAuthor_counts, entityAuthor_IDs, 
                                                            entityAuthor_Aff, last_name, first_name, 
                                                            author, authorID, collabEntity, keys)
                    elif afID.strip() == collabEntity:
                         update_entity_author_counts(entityAuthor_counts, entityAuthor_IDs, 
                                                    entityAuthor_Aff, last_name, first_name, 
                                                    author, authorID, collabEntity, keys)

        # Convertir les dictionnaires en DataFrame
        entity_author_df = pd.DataFrame(list(entityAuthor_counts.items()), columns=['Auteur', 'Nbre de publications'])
            
        # Ajouter les colonne pour les IDs , affiliation des auteurs
        entity_author_df['AU-ID'] = entity_author_df['Auteur'].map(lambda author: ', '.join(entityAuthor_IDs.get(author, [])))
        entity_author_df['AU-ID']  = entity_author_df['AU-ID'].str.replace(', ', '', regex=False)
        entity_author_df['Affiliation'] = entity_author_df['Auteur'].map(lambda author: ', '.join(entityAuthor_Aff.get(author, [])))
        entity_author_df['Affiliation']  = entity_author_df['Affiliation'].str.replace(', ', '', regex=False)
            
        # Trier les résultats
        entity_author_df = entity_author_df.sort_values(by='Nbre de publications', ascending=False).reset_index(drop=True)
        return entity_author_df
    else:
        return
    
def update_entity_author_counts(entityAuthor_counts, entityAuthor_IDs, entityAuthor_Aff, 
                                last_name, first_name, author, authorID, keys, collabEntity = None):
    key = None
    for existing_author in entityAuthor_counts.keys():
        existing_last_name, existing_first_name = existing_author.split(', ')
        if existing_last_name == last_name and existing_first_name.startswith(first_name[0]):
            key = existing_author
            break
    
    if key:
        # Combine counts and keep the author with the longer first name
        entityAuthor_counts[key] += 1
        if len(first_name) > len(key.split(', ')[1]):
            # Update the author ID and affiliation
            entityAuthor_IDs[key] = authorID
            if collabEntity:
                entityAuthor_Aff[key] = getAffiliation(keys, collabEntity)
    else:
        # New author entry
        full_name = f"{last_name}, {first_name}"
        entityAuthor_counts[full_name] = 1
        entityAuthor_IDs[full_name] = authorID
        if collabEntity:
            entityAuthor_Aff[full_name] = getAffiliation(keys, collabEntity)

def load_ETS_profs(console: QPlainTextEdit):
    try : 
        file_name = "INFO.xlsx"
        if os.path.exists(file_name):
            df = pd.read_excel(file_name, sheet_name='Noms_Profs_ETS')
            return df
    except Exception :
        console.append('<p style={}>! Liste des profésseurs ETS non trouvé.</p>'.format(text_style_warning))
        return

def load_ORN(console: QPlainTextEdit):
    try : 
        file_name = "INFO.xlsx"
        if os.path.exists(file_name):
            df = pd.read_excel(file_name, sheet_name='Liste_ORN')
            return df
    except Exception :
        console.append('<p style={}>! Liste des ORN non trouvée.</p>'.format(text_style_warning))
        return
    
def load_UQ(console: QPlainTextEdit):
    try : 
        file_name = "INFO.xlsx"
        if os.path.exists(file_name):
            df = pd.read_excel(file_name, sheet_name='Reseau_UQ')
            return df
    except Exception :
        console.append('<p style={}>! Liste des établissements de l\'UQ non trouvée.</p>'.format(text_style_warning))
        return
def load_ETS(console: QPlainTextEdit):
    try : 
        file_name = "INFO.xlsx"
        if os.path.exists(file_name):
            df = pd.read_excel(file_name, sheet_name='Reseau_ETS')
            return df
    except Exception :
        console.append('<p style={}>! Liste des établissements de l\'ETS non trouvée.</p>'.format(text_style_warning))
        return
    
def add_affiliation_ids_to_list(df: pd.DataFrame, affiliation_list: list, console: QPlainTextEdit):
    try:
        # Vérifie si la colonne 'Affiliation ID' existe dans la DataFrame
        if 'Affiliation ID' in df.columns:
            # Récupère toutes les valeurs de la colonne 'Affiliation ID' et les ajoute à la liste
            affiliation_list.extend([str(aff_id) for aff_id in df['Affiliation ID'].tolist()])
            return affiliation_list
    except Exception:
        console.append('<p style={}>! Erreur lors de la lecture de la colonne Affiliation ID.</p>'.format(text_style_warning))
        return

def findFuzzyMatches(df1: pd.DataFrame, df2: pd.DataFrame, console: QPlainTextEdit):    
    if 'Author' not in df1.columns or 'Nom_prof_ETS' not in df2.columns:
        console.append('<p style={}>! Colonnes Author et/ou Nom_prof_ETS manquantes dans les fichiers</p>'.format(text_style_warning))
        return
    else:
        df1[['Nom', 'Prenom']] = df1['Author'].str.split(', ', expand=True)
        df2[['Nom', 'Prenom']] = df2['Nom_prof_ETS'].str.split(', ', expand=True)
        
        # Remove accents and hyphens, convert to lowercase
        df1['Nom'] = df1['Nom'].fillna('').apply(lambda x: unidecode(x).replace('-', ' ').lower())
        df2['Nom'] = df2['Nom'].fillna('').apply(lambda x: unidecode(x).replace('-', ' ').lower())
        df1['Prenom'] = df1['Prenom'].fillna('').apply(lambda x: unidecode(x).replace('-', ' ').lower())
        df2['Prenom'] = df2['Prenom'].fillna('').apply(lambda x: unidecode(x).replace('-', ' ').lower())
        
        authors = df1[['Nom', 'Prenom', 'Author']].dropna().values.tolist()
        publications = df1.set_index('Author')['Nbre de publications'].to_dict()
        profs = df2[['Nom', 'Prenom', 'Nom_prof_ETS', 'Département']].dropna().values.tolist()
        
        matches = []
        non_matches = []
        fuzzy_matches = []  # To store fuzzy matched authors 
        
        for i, (nom_auteur, prenom_auteur, auteur_complet) in enumerate(authors):
            correspondance_trouvee = False
            for nom_prof, prenom_prof, prof_complet, departement  in profs:
                if nom_auteur == nom_prof and (
                    (len(prenom_auteur) > 2 and prenom_auteur[:2] == prenom_prof[:2]) or 
                    (len(prenom_auteur) <= 2 and prenom_auteur[0] == prenom_prof[0])
                ):
                    matches.append((auteur_complet, prof_complet, departement, publications.get(auteur_complet, 'N/A')))
                    correspondance_trouvee = True
                    break
            if not correspondance_trouvee:
                non_matches.append(auteur_complet)
                
        # Second round with fuzzy matching
        for i, auteur_complet in enumerate(non_matches):
            best_match = None
            highest_ratio = 0
            for nom_prof, prenom_prof, prof_complet, departement in profs:
                ratio = fuzz.ratio(f"{auteur_complet}", f" {prof_complet}")
                if ratio > highest_ratio:
                    highest_ratio = ratio
                    best_match = prof_complet
            if highest_ratio > 80:
                matches.append((auteur_complet, best_match, departement, publications.get(auteur_complet, 'N/A')))
                fuzzy_matches.append(auteur_complet)
                non_matches.remove(auteur_complet)
        
        matches_df = pd.DataFrame(matches, columns=['Auteur', 'Professeur_ETS_correspondant', 'Département', 'Nbre de publications'])
        non_matches_df = pd.DataFrame(non_matches, columns=['Author'])
        non_matches_df['Nbre de publications'] = non_matches_df['Author'].map(publications)
        return matches_df, non_matches_df, fuzzy_matches
def findOthersEtsAffiliations(non_matches_df: pd.DataFrame, all_collabs_df : pd.DataFrame):
        results = []
        if 'affilname' in all_collabs_df.columns and 'Authors' in all_collabs_df.columns:
            affiliations = all_collabs_df['affilname'].dropna().str.split(';')
            authors = all_collabs_df['Authors'].dropna().str.split(';')
            non_matched_authors = non_matches_df['Author']
            nbr_publications = non_matches_df['Nbre de publications']

            for non_matched_author, nbr_publication in zip(non_matched_authors, nbr_publications) :
                gotIt = False
                for affil_list, author_list in zip(affiliations, authors):
                    for affil, author in zip(affil_list, author_list):
                        affil = affil.strip()
                        if author.strip() == non_matched_author:
                            if affil  == 'École de Technologie Supérieure':
                                results.append({
                                'Author': non_matched_author,
                                'Nbre de publications' : nbr_publication
                                    })
                                gotIt = True
                        if gotIt == True:
                            break
                    if gotIt == True:
                        break
                        
        # Conversion des résultats en DataFrame pandas
        other_ets_authors_df = pd.DataFrame(results)
        return other_ets_authors_df

def findCollabCountryAffiliations(non_matches_df: pd.DataFrame, all_collabs_df : pd.DataFrame, collabCountry : str, keys : list):
        results = []
        if 'Authors' in all_collabs_df.columns: 
            authors = all_collabs_df['Authors'].dropna().str.split(';')             
            # countries = all_collabs_df['Countries'].dropna().str.split(';')
            afIDs = all_collabs_df['Authors affiliations'].dropna().str.split(';')
            non_matched_authors = non_matches_df['Author']
            nbr_publications = non_matches_df['Nbre de publications']

            for non_matched_author, nbr_publication in zip(non_matched_authors, nbr_publications) :
                gotIt = False
                for author_list, afID_list in zip( authors, afIDs):
                    for author, afID in zip(author_list, afID_list):
                        if '-' in afID:
                            affs = afID.split('-')
                            for aff in affs:
                                if not aff == '':
                                    affilname, countryCollab = getAffiliationCountry(aff, keys)
                                    if author.strip() == non_matched_author and countryCollab == collabCountry :
                                        results.append({
                                        'Author': non_matched_author,
                                        'Affiliation' : affilname,
                                        'Nbre de publications' : nbr_publication
                                            })
                                        gotIt = True
                                if gotIt == True:
                                    break

                        else : 
                            if not afID == '':
                                affilname, countryCollab = getAffiliationCountry(afID, keys)
                                if author.strip() == non_matched_author and countryCollab == collabCountry:
                                    results.append({
                                    'Author': non_matched_author,
                                    'Affiliation' : affilname,
                                    'Nbre de publications' : nbr_publication
                                        })
                                    gotIt = True
                            if gotIt == True:
                                break
                    if gotIt == True:
                        break
                        
        # Conversion des résultats en DataFrame pandas
        other_authors_df = pd.DataFrame(results)
        return other_authors_df

    
def saveResults(fileName: str, matches_df: pd.DataFrame, other_ets_authors_df : pd.DataFrame, other_authors_df : pd.DataFrame, institutions_df : pd.DataFrame, allResults_df : pd.DataFrame):
        directory = DOCS_PATH[0] + '/' 
        file_path = os.path.join(directory, fileName)
        if not file_path.endswith('.xlsx'):
            file_path += '.xlsx'
                
        # Sort dataframes by 'Nbre de publications' before saving
        matches_df['Nbre de publications'] = matches_df['Nbre de publications'].replace('N/A', 0).astype(int)
        matches_df = matches_df.sort_values(by='Nbre de publications', ascending=False)
        other_ets_authors_df['Nbre de publications'] = other_ets_authors_df['Nbre de publications'].replace('N/A', 0).astype(int)
        other_ets_authors_df = other_ets_authors_df.sort_values(by='Nbre de publications', ascending=False)
        other_authors_df['Nbre de publications'] = other_authors_df['Nbre de publications'].replace('N/A', 0).astype(int)
        other_authors_df = other_authors_df.sort_values(by='Nbre de publications', ascending=False)
            
        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
            matches_df.to_excel(writer, sheet_name='professeurs_ETS', index=False)
            other_ets_authors_df.to_excel(writer, sheet_name='autres_ETS', index=False)
            other_authors_df.to_excel(writer, sheet_name='autres', index=False)
            institutions_df.to_excel(writer, sheet_name='Institutions', index=False)
            allResults_df.to_excel(writer, sheet_name='allResults', index=False)

def highlight_fuzzy_matches(fileName, fuzzy_matches):
        directory = DOCS_PATH[0] + '/' 
        file_path = os.path.join(directory, fileName)
        wb = load_workbook(file_path)
        ws = wb['professeurs_ETS']
        fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for row in ws.iter_rows(min_row=2, max_col=3, max_row=ws.max_row):
            if row[0].value in fuzzy_matches:
                for cell in row:
                    cell.fill = fill
        wb.save(file_path)

# Fonction qui permet d'exporter les données sur le gabarit Excel et d'appeler les
# routines VBA du gabarit
def Excel_collabs_ETS_pays(fileName: str, matches_df: pd.DataFrame, other_ets_authors_df: pd.DataFrame, other_authors_df: pd.DataFrame, institutions_df: pd.DataFrame, allResults_df: pd.DataFrame, fuzzy_matches_df: pd.DataFrame, country : str, debut : str, fin : str, date : str):

    # Remplacer l'extension par .docx
    rapportPath = DOCS_PATH[0] + '\\' + os.path.splitext(fileName)[0] + '.docx'
    gabaritPath = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'GABARITCOLLABS.docx')
    if 'Count' in allResults_df.columns:
            totalDoc = allResults_df['Count']
    infos = {
    'pathToFile': [gabaritPath, rapportPath, 0, 0, 0, 0, 0, 0, 0],
    'totalCount': [allResults_df.shape[0], 0, 0, 0, 0, 0, 0, 0, 0],
    'totalDoc': [totalDoc[0], totalDoc[1], totalDoc[2], totalDoc[3], totalDoc[4], totalDoc[5], totalDoc[6], totalDoc[7], totalDoc[8]],
    'pays': [country, 0, 0, 0, 0, 0, 0, 0, 0],
    'debut':[debut, 0, 0, 0, 0, 0, 0, 0, 0],
    'fin': [fin, 0, 0, 0, 0, 0, 0, 0, 0],
    'date' : [date, 0, 0, 0, 0, 0, 0, 0, 0]
    }
    # Créer le DataFrame
    df_infos = pd.DataFrame(infos)
    # Ouvrir le classeur Excel existant
    template_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'GABARITCOLLABS.xlsm')
    sheet_names = ['Institutions', 'professeurs_ETS', 'autres_ETS', 'autres', 'allResults', 'infos']
    dataframes = [institutions_df, matches_df, other_ets_authors_df, other_authors_df, allResults_df, df_infos]

    # Création de l'objet Excel, et le rendre visible en plein écran lors du processus
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = True
    excel.WindowState = win32.constants.xlMaximized

    try:
        # Vérifier si le fichier Excel est déjà ouvert
        for wb in excel.Workbooks:
            if wb.FullName == template_file:
                wb.Close(False)  # Fermer le classeur sans enregistrer les modifications

        # Ouverture du fichier Excel
        workbook = excel.Workbooks.Open(template_file)
        workbook.Visible = True  # Rendre le classeur visible
        workbook.WindowState = win32.constants.xlMaximized  # Mettre le classeur en plein écran 

        # Enregistrer sous le nouveau nom
        workbook.SaveAs(os.path.abspath(DOCS_PATH[0] + '/' + fileName), FileFormat=52)

        # Écrire les données des DataFrames dans les feuilles de calcul
        for sheet_name, df in zip(sheet_names, dataframes):
            worksheet = workbook.Worksheets(sheet_name)
            for i, col in enumerate(df.columns):
                worksheet.Cells(1, i + 1).Value = col
            for i, row in df.iterrows():
                for j, value in enumerate(row):
                    worksheet.Cells(i + 2, j + 1).Value = value

        # Mise en surbrillance des lignes correspondantes dans 'professeurs_ETS'
        professeurs_sheet = workbook.Worksheets('professeurs_ETS')
        yellow = 65535  # Code couleur pour le jaune en format RGB   
        for row in range(2, professeurs_sheet.UsedRange.Rows.Count + 1):
            cell_value = professeurs_sheet.Cells(row, 1).Value  # Assurez-vous que les noms des professeurs sont en colonne 1
            if cell_value in fuzzy_matches_df:
                for col in range(1, professeurs_sheet.UsedRange.Columns.Count + 1):
                    professeurs_sheet.Cells(row, col).Interior.Color = yellow  # Jaune

        # Mettre la fenêtre en premier plan
        try:
            win32gui.SetForegroundWindow(win32gui.FindWindow(None, workbook.Name + " - Excel"))
        except:
            win32gui.SetForegroundWindow(win32gui.FindWindow(None, workbook.Name.split('.')[0] + " - Excel"))

        # Activer la feuille 'professeurs_ETS'
        professeurs_sheet.Activate()

        #enregistrer les modifications
        workbook.Save()
        # Déverrouiller le fichier pour modification
        workbook.Protect(Structure=False, Windows=False)
        # Appel de la procédure VBA
        nom_module = 'Module1'
        nom_procedure = 'Workbook_Open'
        excel.Run(f'{nom_module}.{nom_procedure}')

        # Mettre à jour l'affichage pour s'assurer que l'utilisateur peut voir et interagir avec le classeur
        excel.ScreenUpdating = True
        excel.Interactive = True
        workbook.Activate()

    except Exception as e:
        return f"Une erreur s'est produite : {e}"
    return excel, workbook
    


def Excel_autes_collabs(fileName: str, matches_df: pd.DataFrame, other_ets_authors_df: pd.DataFrame, other_authors_df: pd.DataFrame, institutions_df: pd.DataFrame, allResults_df: pd.DataFrame, fuzzy_matches_df: pd.DataFrame):

    # Remplacer l'extension par .docx
    rapportPath = DOCS_PATH[0] + '\\' + os.path.splitext(fileName)[0] + '.docx'
    gabaritPath = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'GABARITCOLLABS.docx')
    paths = {
    'pathToFile': [gabaritPath, rapportPath]
    }
    # Créer le DataFrame
    df_paths = pd.DataFrame(paths)
    # Ouvrir le classeur Excel existant
    template_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'GABARITCOLLABS.xlsm')
    sheet_names = ['Institutions', 'professeurs_ETS', 'autres_ETS', 'autres', 'allResults', 'paths']
    dataframes = [institutions_df, matches_df, other_ets_authors_df, other_authors_df, allResults_df, df_paths]

    # Création de l'objet Excel, et le rendre visible en plein écran lors du processus
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = True
    excel.WindowState = win32.constants.xlMaximized

    try:
        # Vérifier si le fichier Excel est déjà ouvert
        for wb in excel.Workbooks:
            if wb.FullName == template_file:
                wb.Close(False)  # Fermer le classeur sans enregistrer les modifications

        # Ouverture du fichier Excel
        workbook = excel.Workbooks.Open(template_file)
        workbook.Visible = True  # Rendre le classeur visible
        workbook.WindowState = win32.constants.xlMaximized  # Mettre le classeur en plein écran 

        # Enregistrer sous le nouveau nom
        workbook.SaveAs(os.path.abspath(DOCS_PATH[0] + '/' + fileName), FileFormat=52)

        # Écrire les données des DataFrames dans les feuilles de calcul
        for sheet_name, df in zip(sheet_names, dataframes):
            worksheet = workbook.Worksheets(sheet_name)
            for i, col in enumerate(df.columns):
                worksheet.Cells(1, i + 1).Value = col
            for i, row in df.iterrows():
                for j, value in enumerate(row):
                    worksheet.Cells(i + 2, j + 1).Value = value

        # Mise en surbrillance des lignes correspondantes dans 'professeurs_ETS'
        professeurs_sheet = workbook.Worksheets('professeurs_ETS')
        yellow = 65535  # Code couleur pour le jaune en format RGB   
        for row in range(2, professeurs_sheet.UsedRange.Rows.Count + 1):
            cell_value = professeurs_sheet.Cells(row, 1).Value  # Assurez-vous que les noms des professeurs sont en colonne 1
            if cell_value in fuzzy_matches_df:
                for col in range(1, professeurs_sheet.UsedRange.Columns.Count + 1):
                    professeurs_sheet.Cells(row, col).Interior.Color = yellow  # Jaune

        # Mettre la fenêtre en premier plan
        try:
            win32gui.SetForegroundWindow(win32gui.FindWindow(None, workbook.Name + " - Excel"))
        except:
            win32gui.SetForegroundWindow(win32gui.FindWindow(None, workbook.Name.split('.')[0] + " - Excel"))

        # Activer la feuille 'professeurs_ETS'
        professeurs_sheet.Activate()

        #enregistrer les modifications
        workbook.Save()
        # Déverrouiller le fichier pour modification
        workbook.Protect(Structure=False, Windows=False)
        # Appel de la procédure VBA
        nom_module = 'Module1'
        nom_procedure = 'Workbook_Open'
        excel.Run(f'{nom_module}.{nom_procedure}')

        # Mettre à jour l'affichage pour s'assurer que l'utilisateur peut voir et interagir avec le classeur
        excel.ScreenUpdating = True
        excel.Interactive = True
        workbook.Activate()

    except Exception as e:
        print(f"Une erreur s'est produite : {e}")
    return excel, workbook

def saveInter(fileName :str, dfAllResults :pd.DataFrame, dfAuteurs :pd.DataFrame = None, dfAuteursA :pd.DataFrame = None, dfAuteursB :pd.DataFrame = None, dfInstitutions :pd.DataFrame = None):
# def saveInter(dfAllResults :pd.DataFrame, fileName :str):
    directory = DOCS_PATH[0] + '/' 
    file_path = os.path.join(directory, fileName)
    if not file_path.endswith('.xlsx'):
        file_path += '.xlsx'
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # Écrire dfAllResults en premier
        dfAllResults.to_excel(writer, sheet_name='allResults', index=False)
        
        # Écrire les autres DataFrames si elles ne sont pas None ou vides
        if dfAuteurs is not None and not dfAuteurs.empty:
            dfAuteurs.to_excel(writer, sheet_name='Liste des auteurs', index=False)
        
        if dfAuteursA is not None and not dfAuteursA.empty:
            dfAuteursA.to_excel(writer, sheet_name='Auteurs entité A', index=False)
        
        if dfAuteursB is not None and not dfAuteursB.empty:
            dfAuteursB.to_excel(writer, sheet_name='Auteurs entité B', index=False)
        
        if dfInstitutions is not None and not dfInstitutions.empty:
            dfInstitutions.to_excel(writer, sheet_name='Institutions entité B', index=False)

    return

def getAffiliation(InstitutionId: str, keys: list):
    query_entity = f'AF-ID({InstitutionId})'
    search = AffiliationSearch(query=query_entity, api_key= keys[0], token= keys[1])
    # Vérification des résultats
    if search.affiliations is None:
        return 'NONE'
    else:
        for element in search.affiliations:
            affiliation = f"{element.name}"
        if affiliation:
            return affiliation
        else: return 'NONE'
def getAffiliationCountry(InstitutionId: str, keys: list):
    query_entity = f'AF-ID({InstitutionId})'
    search = AffiliationSearch(query=query_entity, api_key= keys[0], token= keys[1])
    # Vérification des résultats
    if search.affiliations is None:
        return 'NONE'
    else:
        for element in search.affiliations:
            affiliationCountry = f"{element.country}"
            affiliation = f"{element.name}"
        if affiliation:
            return affiliation, affiliationCountry
        else: return 'NONE'
def getAuthorORCID(authorId: str, keys: list):
    query_entity = f'AU-ID({authorId})'
    search = AuthorSearch(query=query_entity, api_key= keys[0], token= keys[1])
    # Vérification des résultats
    if search.authors is None:
        return 'NONE'
    else:
        for element in search.authors:
            orcid = element.orcid
        if orcid:
            return orcid
        else: return 'NONE'

def getAbstract(EID: str, keys: list):
    try : 
        search = AbstractRetrieval(identifier=EID, api_key= keys[0], token= keys[1])
        # Vérification des résultats
        if search.abstract is None:
            if search.description is None:
                return 'NOT FOUND'
            else:
                return  search.description
            
        else:
            return search.abstract
    except:
        return 'NONE'